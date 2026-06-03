#!/usr/bin/env python
"""把 CSV 转成带格式的 Excel 报表。"""
import argparse
import csv
import json
import sys
from pathlib import Path


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--csv_path", required=True)
    p.add_argument("--output_path", required=True)
    p.add_argument("--title", default=None)
    args = p.parse_args()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[error] openpyxl 未安装，请: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    csv_path = Path(args.csv_path)
    if not csv_path.exists():
        print(f"[error] CSV 不存在: {csv_path}", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    row_offset = 1
    if args.title:
        ws.cell(row=1, column=1, value=args.title).font = Font(bold=True, size=14)
        row_offset = 3

    with open(csv_path, encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)

    # 写表头
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    if rows:
        for col_idx, header in enumerate(rows[0], start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    # 写数据
    for row_idx, row in enumerate(rows[1:], start=row_offset + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 简单自动列宽
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    out_path = Path(args.output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    result = {
        "output_path": str(out_path),
        "size_bytes": out_path.stat().st_size,
        "rows_written": len(rows),
        "columns_count": len(rows[0]) if rows else 0,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
